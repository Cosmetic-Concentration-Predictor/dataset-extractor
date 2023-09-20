import os
import pandas as pd
import openpyxl
import xlrd
import re


class ExcelProcessor:
    def __init__(self, input_folder, output_csv, output_txt):
        self.input_folder = input_folder
        self.output_csv = output_csv
        self.output_txt = open(output_txt, "w")

    def process_excel_files(self):
        result_df = pd.DataFrame(columns=["acronym", "code", "concentration"])

        for filename in os.listdir(self.input_folder):
            print(filename)
            if filename.lower().endswith(".xlsx"):
                file_path = os.path.join(self.input_folder, filename)
                result_df = pd.concat(
                    [result_df, self.process_xlsx(file_path)], ignore_index=True
                )
            elif filename.lower().endswith(".xls"):
                file_path = os.path.join(self.input_folder, filename)
                result_df = pd.concat(
                    [result_df, self.process_xls(file_path)], ignore_index=True
                )
            # else:
            #     print(filename)

        result_df.to_csv(self.output_csv, index=False)

    def process_xlsx(self, file_path):
        # Abre o arquivo Excel usando openpyxl
        wb = openpyxl.load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            start_row, start_col, end_row, end_col = self.find_table_range(sheet)

            if start_row is not None and start_col is not None:
                # Encontrou a tabela, agora podemos processar
                table_data = []
                for row_num in range(start_row + 1, end_row + 1):
                    if "PHASE" in str(
                        sheet.cell(row=row_num, column=start_col).value
                    ) or " " in str(sheet.cell(row=row_num, column=start_col).value):
                        continue

                    row_data = [
                        sheet.cell(row=row_num, column=start_col).value,
                        sheet.cell(row=row_num, column=start_col + 2).value,
                    ]
                    table_data.append(row_data)

                # Transforma a tabela em DataFrame
                table_df = pd.DataFrame(table_data, columns=["Code", "Pour 1"])

                # Remove o símbolo de porcentagem e converte para float se aplicável
                table_df["Pour 1"] = (
                    table_df["Pour 1"]
                    .replace("%", "", regex=True)
                    .apply(pd.to_numeric, errors="coerce")
                )

                # Adiciona as letras e números do 'Code' às colunas 'acronym' e 'code'
                table_df["acronym"] = table_df["Code"].str.extract("([a-zA-Z]+)")
                table_df["code"] = table_df["Code"].str.extract("(\d+)")
                table_df["concentration"] = table_df[table_df["Pour 1"] > 0.0]["Pour 1"]

                table_df.drop("Code", axis=1, inplace=True)
                table_df.drop("Pour 1", axis=1, inplace=True)
                table_df.dropna(inplace=True)

                table_df["acronym"] = table_df["acronym"].apply(
                    lambda x: re.sub(
                        "^(ME|MT|AMT|LABOMT|LABOME|LABOAMT|LABO|LAB)", "", x.upper()
                    )
                )

                self.output_txt.write(f"{file_path}: {sheet.name}\n{table_df}\n\n")

                # Concatena os resultados com o DataFrame principal
                result_df = pd.concat([result_df, table_df], ignore_index=True)
            # else:
            #     print(file_path)

        # Fecha o arquivo Excel
        wb.close()

        # Retorna o DataFrame atualizado
        return result_df

    def process_xls(self, file_path):
        result_df = None
        wb = xlrd.open_workbook(file_path)

        for sheet in wb.sheets():
            start_row, start_col, end_row, end_col = self.find_table_range_xls(sheet)

            if start_row is not None and start_col is not None:
                table_data = []
                for row_num in range(start_row, end_row):
                    if (
                        "PHASE" not in str(sheet.cell_value(row_num, start_col))
                        or " " not in str(sheet.cell_value(row_num, start_col))
                        or "Code" not in str(sheet.cell_value(row_num, start_col))
                        or "Pour" not in str(sheet.cell_value(row_num, start_col))
                    ):
                        row_data = [
                            sheet.cell_value(row_num, start_col),
                            sheet.cell_value(row_num, start_col + 2),
                        ]
                        table_data.append(row_data)
                    else:
                        continue

                table_df = pd.DataFrame(table_data, columns=["Code", "Pour 1"])
                table_df.dropna(inplace=True)

                table_df["Pour 1"] = (
                    table_df["Pour 1"]
                    .replace("%", "", regex=True)
                    .apply(pd.to_numeric, errors="coerce")
                )

                table_df["acronym"] = table_df["Code"].str.extract(
                    r"([A-Za-z]+\s?(?=\d)(?:\d+[A-Za-z]+)?)"
                )
                table_df["code"] = table_df["Code"].str.extract(
                    r"(\s?\d*(?!.*[A-Za-z]))"
                )

                table_df["concentration"] = table_df["Pour 1"].astype(float)

                table_df.drop("Code", axis=1, inplace=True)
                table_df.drop("Pour 1", axis=1, inplace=True)
                table_df["acronym"] = table_df[table_df["acronym"].str.len() > 1][
                    "acronym"
                ]
                table_df["acronym"] = table_df["acronym"].str.rstrip()
                table_df.dropna(inplace=True)

                df_sum = table_df["concentration"].sum()
                if df_sum > 1.8:
                    table_df["concentration"] = table_df["concentration"].div(100)

                table_df["acronym"] = table_df["acronym"].apply(
                    lambda x: re.sub(
                        "^(§|MY|ME|MT|AMT|LABOMT|LABOME|LABOAMT|LABO|LAB)",
                        "",
                        x.upper(),
                    )
                )

                try:
                    table_df["code"] = table_df["code"].astype(int)
                except ValueError:
                    print(f"{file_path} - {sheet.name}")
                    continue

                self.output_txt.write(f"{file_path}: {sheet.name}\n{table_df}\n\n")

                result_df = pd.concat([result_df, table_df], ignore_index=True)
                result_df.dropna(inplace=True)
            # else:
            #     print(f"{file_path}: {sheet.name}")

        return result_df

    def find_table_range(self, sheet):
        for row_num in range(1, sheet.max_row + 1):
            for col_num in range(1, sheet.max_column - 2):
                if (
                    (
                        sheet.cell(row=row_num, column=col_num).value == "Code"
                        and sheet.cell(row=row_num, column=col_num).value
                        != "Code action"
                    )
                    and (
                        sheet.cell(row=row_num, column=col_num + 2).value == "Pour 1"
                        or sheet.cell_value(row_num, col_num + 2) == "%global"
                    )
                ) or (
                    (
                        sheet.cell(row=row_num, column=col_num).value == ""
                        or sheet.cell(row=row_num, column=col_num).value == "N°MP"
                    )
                    and sheet.cell(row=row_num, column=col_num + 2).value == "%"
                ):
                    start_row, start_col = row_num, col_num
                    end_row, end_col = sheet.max_row, col_num + 2
                    return start_row, start_col, end_row, end_col

        return None, None, None, None

    def find_table_range_xls(self, sheet):
        for row_num in range(1, sheet.nrows):
            for col_num in range(0, sheet.ncols - 2):
                if (
                    (
                        str(sheet.cell_value(row_num, col_num)).lower() == "code"
                        and str(sheet.cell_value(row_num, col_num)).lower()
                        != "code action"
                    )
                    and (
                        sheet.cell_value(row_num, col_num + 2) == "Pour 1"
                        or sheet.cell_value(row_num, col_num + 2) == "%global"
                    )
                ) or (
                    (
                        sheet.cell_value(row_num, col_num) == ""
                        or sheet.cell_value(row_num, col_num) == "N°MP"
                    )
                    and sheet.cell_value(row_num, col_num + 2) == "%"
                ):
                    start_row, start_col = row_num, col_num
                    end_row, end_col = sheet.nrows, col_num + 2
                    return start_row, start_col, end_row, end_col

        return None, None, None, None
